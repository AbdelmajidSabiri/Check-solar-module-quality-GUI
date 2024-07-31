import pyvisa
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

class Bkp8600(object):
    def __init__(self, resource=None):
        self.rm = pyvisa.ResourceManager()
        self.instr = None 
        self.instrument_found = False
        if resource:
            try:
                self.instr = self.rm.open_resource(resource, timeout=10000)
                self.instrument_found = True
            except pyvisa.VisaIOError:
                print(f"Error: Unable to open resource {resource}.")
        else:
            instruments = self.rm.list_resources()
            for r in instruments:
                try:
                    instr = self.rm.open_resource(r)
                    if instr.query("*IDN?").startswith("B&K Precision"):
                        self.instr = instr
                        self.instrument_found = True
                        break
                except pyvisa.VisaIOError:
                    pass
            if not self.instrument_found:
                print("No BK Precision instrument found.")

    def get_description(self):
        if self.instrument_found:
            return self.instr.query("*IDN?")

    def get_current(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:CURRent?"))

    def get_voltage(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:VOLTage?"))
    
    def get_resistance(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:RESistance?"))

    def get_power(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:POWer?"))

    def initialize(self):
        if self.instrument_found:
            self.instr.write("SYSTem:REMote")
            self.instr.write("INPut OFF")
            self.instr.write("*RST")
            self.instr.write("*CLS")
            self.instr.write("*SRE 0")
            self.instr.write("*ESE 0")

    def set_current(self, current):
        if self.instrument_found:
            self.instr.write("INPut ON")
            self.instr.write("FUNC CURRent")
            self.instr.write(f"CURRent {current}")

    
    def set_CV(self,current_limit):
        if self.instrument_found:
            self.instr.write("*RST")
            self.instr.write("FUNC VOLTage")  # Set the function mode to Constant Voltage (CV)
            self.instr.write(f"CURR:LIMIT {current_limit}") 
    
    def set_voltage(self,voltage) :
        if self.instrument_found:
            self.instr.write("INPut ON")
            self.instr.write(f"VOLT {voltage}")



    def reset_to_manual(self):
        if self.instrument_found:
            self.instr.write("INPut OFF")
            self.instr.write("SYSTem:LOCal")

        
if __name__ == '__main__':
    try:
        last = Bkp8600()

        last.initialize()

        # Perform operations (e.g., setting current and measuring values)
        for n in range(0, 200):
            current_value = n / 100.0
            print(f"Setting current to: {current_value} A")
            last.set_current(current_value)
            time.sleep(1)
            print(f"Current: {last.get_current()} A")
            print(f"Voltage: {last.get_voltage()} V")



    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Reset to local mode and close the connection
        last.reset_to_manual()
        del last


def CollectData(date, time, serial_number, max_power = 0, Impp=0, Vmpp=0, Voc=0, Isc=0, FF = 0, Grade = "A"):
    wb = load_workbook('output.xlsx')


    # Select the worksheet
    sheet = wb['Sheet2']

    last_id = sheet.cell(row=sheet.max_row, column=1).value  # Assuming ID is in column 1
    last_module_number = sheet.cell(row=sheet.max_row, column=2).value 


    if last_id is not None:
        next_id = int(last_id) + 1
    else:
        next_id = 1 


    if last_module_number is not None:
        next_module_number = int(last_module_number) + 1
    else:
        next_module_number = 312  

    # Append data to the worksheet
    row = ( next_id,                        #ID
            next_module_number,             #Module Number
            0,                              #Test Result
            max_power,                      #Pmpp
            Impp,                           #Imp
            Vmpp,                           #Vmp
            Voc,                            #Uoc
            Isc,                            #Isc
            FF,                             #Fil factor
            Grade,                          #Grade
            0,                              #Recurrence
            0,                              #Temperature Ambient
            0,                              #Temperature Lamps
            31.0654,                        #Pmpp Reference
            0,                              #Pmpp Deviation
            9.9332,                         #Uoc Reference
            0,                              #Uoc Deviation
            4.3456,                         #Isc Reference
            0,                              #Isc Deviation
            0,                              #Temperature Ambient Reference
            2,                              #Temperature Lamps Reference
            0,                              #Reference Number
            date,                           #Date
            time,                           #Time
            serial_number                   #Serial Number
            )

    sheet.append(row)

    # Save the workbook
    wb.save('output.xlsx')