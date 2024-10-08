from pathlib import Path
from tkinter import Tk, Canvas, Entry, Button, PhotoImage,ttk, DoubleVar,Label,StringVar, IntVar,messagebox,simpledialog
import tkinter as tk
from tkinter.font import Font
import customtkinter as ctk
import pyvisa
# import pyvisa_py
import serial
import serial.tools.list_ports
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
import numpy as np
import sys,time
import os
import math
import random
sys.path.append('C:\\Users\\dell\\GUI-read-Bk-precision-data')
from PIL import Image, ImageTk
from datetime import datetime
from openpyxl import load_workbook, Workbook


class Bkp8600(object):
    def __init__(self, resource=None):
        self.rm = pyvisa.ResourceManager()
        self.instr = None 
        self.instrument_found = False
        if resource:
            try:
                self.instr = self.rm.open_resource(resource, timeout=10000)
                self.instrument_found = True
            except pyvisa.VisaIOError:
                print(f"Error: Unable to open resource {resource}.")
        else:
            instruments = self.rm.list_resources()
            for r in instruments:
                try:
                    instr = self.rm.open_resource(r)
                    if instr.query("*IDN?").startswith("B&K Precision"):
                        self.instr = instr
                        self.instrument_found = True
                        break
                except pyvisa.VisaIOError:
                    pass
            if not self.instrument_found:
                print("No BK Precision instrument found.")

    #Functions to communicate with Bk precision 8600
    def get_description(self):
        if self.instrument_found:
            return self.instr.query("*IDN?")

    def get_current(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:CURRent?"))

    def get_voltage(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:VOLTage?"))
    
    def get_resistance(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:RESistance?"))

    def get_power(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:POWer?"))

    def initialize(self):
        if self.instrument_found:
            self.instr.write("SYSTem:REMote")
            self.instr.write("INPut OFF")
            self.instr.write("*RST")
            self.instr.write("*CLS")
            self.instr.write("*SRE 0")
            self.instr.write("*ESE 0")

    def set_current(self, current):
        if self.instrument_found:
            self.instr.write("INPut ON")
            self.instr.write("FUNC CURRent")
            self.instr.write(f"CURRent {current}")
    
    def set_CV(self,current_limit):
        if self.instrument_found:
            self.instr.write("*RST")
            self.instr.write("FUNC VOLTage")  # Set the function mode to Constant Voltage (CV)
            self.instr.write(f"CURR:LIMIT {current_limit}") 
    
    def set_voltage(self,voltage) :
        if self.instrument_found:
            self.instr.write("INPut ON")
            self.instr.write(f"VOLT {voltage}")

    def reset_to_manual(self):
        if self.instrument_found:
            self.instr.write("INPut OFF")
            self.instr.write("SYSTem:LOCal")


class ArduinoCmd:
    def __init__(self, port='COM3'):
        self.arduino = None
        self.arduino_found = False

        self.try_connect(port)



    # Fucntion to connect to arduino with a specifique port
    def try_connect(self, port):
        try:
            # Attempt to open the specified COM port
            self.arduino = serial.Serial(port, 9600, timeout=1)
            self.arduino_found = True
        except serial.SerialException:
            self.arduino = None
            self.arduino_found = False

    # Function to turn light on
    def turn_light_on(self):
        if self.arduino :
            self.arduino.write(b'1')

    # Function to turn light off
    def turn_light_off(self):
        if self.arduino :
            self.arduino.write(b'0')

    # Function to update temperature that cams from a temperature sensor
    def update_temperature(self):
        self.arduino.write(b'T')
        if self.arduino.in_waiting > 0:
            temp_data = self.arduino.readline().decode().strip()
            if temp_data:
                self.temp_var.set(temp_data)

    # Function to close connection with arduino
    def close(self):
        if self.arduino is not None:
            self.arduino.close()

class GUI:
    def __init__(self):

        self.bk_device = Bkp8600()
        self.bk_device.initialize()

        self.Arduino = ArduinoCmd()

        self.data_list_current_measured = []
        self.data_list_voltage_measured = []
        self.data_list_power = []
        self.options_list = ["Profile 1", "Profile 2","Add Profile"]
        
        self.max_power = 0.0
        self.MPP = {"Vmpp" : 0 , "Impp" : 0}
        self.progress = 0
        self.running = False

        self.window = ctk.CTk()
        self.window.geometry("1420x800")
        self.window.title("Agamine Solar LAMINATE TESTING V 1.0")
        self.window.configure(fg_color = "#D7E1E7")



        # Create variables for displaying max power and other parameters
        self.Current_var = DoubleVar(value=0.00)
        self.Voltage_var =  DoubleVar(value=0.00)
        self.Power_var = DoubleVar(value=0.00)
        self.max_power_var = DoubleVar(value=0.00)
        self.Vmpp_var = DoubleVar(value= 0.00)
        self.Impp_var = DoubleVar(value= 0.00)
        self.Isc_var = DoubleVar(value= 0.00)
        self.Voc_var = DoubleVar(value= 0.00)
        self.FF_var = DoubleVar(value= 0.00)
        self.temp_var = DoubleVar(value=25)
        self.grade_var = StringVar(value="?")
        self.recurrence_var = IntVar(value = 0)
        self.serial_num_var = StringVar()
        self.result_var = StringVar(value = "Passe")

        self.mode_var = tk.StringVar(value="CV")  # Default selection

        self.profiles = [{  "name" : "Profile 1",
                            "start voltage" : DoubleVar(value = 0.1),
                            "stop voltage" : DoubleVar(value = 12),
                            "step size voltage" : DoubleVar(value = 0.1),
                            "dwell time voltage" : DoubleVar(value = 0.1),
                            "start current" : DoubleVar(value = 0.1),
                            "stop current" : DoubleVar(value = 15),
                            "step size current" : DoubleVar(value = 0.1),
                            "dwell time current" : DoubleVar(value = 0.1),
                            "current limit" : DoubleVar(value = 12),
                            "voltage limit" : DoubleVar(value = 20),
                            "power limit" : DoubleVar(value = 240),
                            "temperature limit" : DoubleVar(value = 30),
                            "current resolution" : DoubleVar(value = 0.1),
                            "voltage resolution" : DoubleVar(value = 0.1),
                            "active profile" : 1,
                            },
                            {"name" : "Profile 2",
                            "start voltage" : DoubleVar(value = 0.1),
                            "stop voltage" : DoubleVar(value = 15),
                            "step size voltage" : DoubleVar(value = 0.2),
                            "dwell time voltage" : DoubleVar(value = 0.1),
                            "start current" : DoubleVar(value = 0.1),
                            "stop current" : DoubleVar(value = 50),
                            "step size current" : DoubleVar(value = 0.1),
                            "dwell time current" : DoubleVar(value = 0.1),
                            "current limit" : DoubleVar(value = 12),
                            "voltage limit" : DoubleVar(value = 20),
                            "power limit" : DoubleVar(value = 900),
                            "temperature limit" : DoubleVar(value = 30),
                            "current resolution" : DoubleVar(value = 0.1),
                            "voltage resolution" : DoubleVar(value = 0.1),
                            "active profile" : 0,  
                            }]

        self.selected_option = tk.StringVar(value="Profile 1")
        self.selected_index = 0
        self.selected_profile = self.profiles[self.selected_index]
        self.working_profile = self.profiles[self.selected_index]
        

        self.columns = ["Serial Num","Results","Date", "Test Time",
                "Mpp", "Imp", "Vmp","Isc",
                "Voc","FF", 
                "Grade","Recurrence","Profile"]
        self.TableData = pd.DataFrame(columns=self.columns)



        self.Current_formated = StringVar(value="0.00")
        self.Voltage_formated =  StringVar(value="0.00")
        self.Power_formated = StringVar(value="0.00")
        self.max_power_formated = StringVar(value="0.00")
        self.Vmpp_formated = StringVar(value="0.00")
        self.Impp_formated = StringVar(value="0.00")
        self.Isc_formated = StringVar(value="0.00")
        self.Voc_formated = StringVar(value="0.00")
        self.FF_formated = StringVar(value="0.00")


        self.left_frame = ctk.CTkFrame(master=self.window, width=200, corner_radius=0, fg_color='white')
        self.left_frame.pack(side="left", fill="y")

        # Load images
        self.image_dashboard_ON_img = PhotoImage(file=self.get_image_path("dashboard_ON.png"))
        self.image_dashboard_OFF_img = PhotoImage(file=self.get_image_path("dashboard_OFF.png"))

        self.image_bk_profiles_OFF_img = PhotoImage(file=self.get_image_path("Bk_profiles_OFF.png"))
        self.image_bk_profiles_ON_img = PhotoImage(file=self.get_image_path("Bk_profiles_ON.png"))

        self.fram_indicator_img = PhotoImage(file=self.get_image_path("Fram_indicator.png"))
        self.vertical_line_img = PhotoImage(file=self.get_image_path("Vertical_line.png"))

        # Create buttons with images
        self.dashboard_button = ctk.CTkButton(master=self.left_frame, 
                                                image=self.image_dashboard_ON_img, 
                                                text="DASHBOARD", 
                                                compound="top", 
                                                command=self.show_dashboard, 
                                                fg_color='white', 
                                                text_color='#0000ff', 
                                                font=("Arial Rounded MT Bold",12), 
                                                hover_color="white")
        self.dashboard_button.pack(pady=10, padx=0, anchor='w')

        self.bk_profiles_button = ctk.CTkButton(master=self.left_frame,
                                                 image=self.image_bk_profiles_OFF_img, 
                                                 text="BK PROFILES", 
                                                 compound="top", 
                                                 command=self.show_bk_profiles, 
                                                 fg_color='white', 
                                                 text_color='#b7b7fc', 
                                                 font=("Arial Rounded MT Bold",12), 
                                                 hover_color="white")
        self.bk_profiles_button.pack(pady=10, padx=0, anchor='w')

        self.about_button = ctk.CTkButton(  master=self.left_frame, 
                                            text="About",
                                            command=self.About,
                                            fg_color='white', 
                                            text_color='#0000ff',
                                            font=("Arial",13),
                                            hover_color="white",
                                            )
        self.about_button.pack(pady=(560,0), padx=0, anchor='w')


        # Create a frame for the main content
        self.main_frame = ctk.CTkFrame(master=self.window, fg_color='#D7E1E7')
        self.main_frame.pack(side="right", expand=True, fill="both", padx=0, pady=0)

        self.dashboard_frame = ctk.CTkFrame(master=self.main_frame, fg_color='#D7E1E7')
        self.bk_profiles_frame = ctk.CTkFrame(master=self.main_frame, fg_color='#D7E1E7')

        # Initial display
        self.dashboard_frame.pack(fill="both", expand=True)

        # Setup content of dashboard
        self.setup_dashboard_content()
        self.setup_bk_profiles_content()

        # Start update loop for maximum power and time

        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)  # Handle window closing event
        self.window.mainloop()



    # Function to get relative path of images
    def get_image_path(self, image_name):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(script_dir, 'images', image_name)
        return image_path

    # Funtion to Get data from Bk-precision
    def get_data(self):
        try:
            measured_current = self.bk_device.get_current()
            voltage = self.bk_device.get_voltage()
            if measured_current and voltage :
                return measured_current, voltage
            else :
                return 0.00,0.00
        except Exception as e:
            print(f"Error retrieving data: {e}")
            return 0.00, 0.00

    # Method to get Serial Number
    def get_serialNum(self):
        return int(self.serial_num_var.get())

    def get_month_folder(self):
        current_month = datetime.now().strftime("%B")

        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__)) 

        data_folder = os.path.join(base_dir, 'Data')
        if not os.path.exists(data_folder):
            os.makedirs(data_folder)

        month_folder = os.path.join(data_folder, current_month)

        if not os.path.exists(month_folder):
            os.makedirs(month_folder)

        return month_folder

    # Funtion to Get Max power
    def update_max_power(self) :
        if self.Power_var.get() > self.max_power_var.get():
            self.max_power_var.set(self.Power_var.get())
            self.Impp_var.set(self.Current_var.get())
            self.Vmpp_var.set(self.Voltage_var.get())

            self.max_power_formated.set(f"{self.max_power_var.get():.2f}")
            self.Impp_formated.set(f"{self.Impp_var.get():.2f}")
            self.Vmpp_formated.set(f"{self.Vmpp_var.get():.2f}")

    # Function to update data
    def update_data(self) :
        if self.running :
            self.update_max_power()
            self.calculate_isc_voc()
 
   # Calculate Isc and Voc
    def calculate_isc_voc(self):

        if self.data_list_current_measured:
            self.Isc_var.set(max(self.data_list_current_measured))
        else :
            self.Isc_var.set(0.00)

        self.Isc_formated.set(f"{self.Isc_var.get():.2f}")


        if self.data_list_voltage_measured:
            self.Voc_var.set(max(self.data_list_voltage_measured))
        else :
            self.Voc_var.set(0.00)

        self.Voc_formated.set(f"{self.Voc_var.get():.2f}")
  
    # Function to calculate fill factor
    def calculate_FF_Grade(self) :
        isc = self.Isc_var.get()
        voc = self.Voc_var.get()
        maxPower = self.max_power_var.get()

        if isc and voc and maxPower :
            FF = (maxPower/(isc * voc)) * 100
        else :
            FF = 0.00

        self.FF_var.set(FF)
        self.FF_formated.set(f"{self.FF_var.get():.2f}")

        self.grade_var.set("C")

        if self.FF_var.get() < 50 :
            self.grade_var.set("C")
        elif self.FF_var.get() < 60 :
            self.grade_var.set("B")
        else :
            self.grade_var.set("A")  

    # Function to calculate recurrence
    def calculate_recurrence(self):

        self.create_excel_file_if_not_exists()

        month_folder = self.get_month_folder()
        excel_file_path = os.path.join(month_folder, 'Data.xlsx')

        df = pd.read_excel(excel_file_path)

        serial_number = self.get_serialNum()
        self.recurrence_var.set((df["Serial Number"] == serial_number).sum() + 1)
  
    # Function to find result of test
    def find_result(self) :
        if self.FF_var.get() < 60 :
            self.result_var.set("Fail")
        else :
            self.result_var.set("Passe")

    # Function to set design of plot
    def configure_plot(self):
        self.ax.set_facecolor('#e4e4e4')
        self.ax.spines['top'].set_visible(False)
        self.ax.spines['right'].set_visible(False)

        self.ax.set_xlabel("Voltage (V)", fontsize=8, fontweight='bold')
        self.ax.set_ylabel("Current (A) / Power (W)", fontsize=8, fontweight='bold')

        self.ax.tick_params(axis='x', colors='#000000')
        self.ax.tick_params(axis='y', colors='#000000')

        self.ax.xaxis.label.set_color('#000000')
        self.ax.yaxis.label.set_color('#000000')

        self.ax.grid(True, color='#3A3A3A', linestyle='--', linewidth=0.5)

    # Function to animate chart
    def animate_chart(self,current_measured,voltage_measured):
        if self.running :
            
            self.ax.clear()
            self.configure_plot()
                    
            # Plot the current data
            self.ax.plot(self.data_list_voltage_measured,self.data_list_current_measured,color='Blue', linewidth=2, label='I-V curve')        
            self.ax.plot(self.data_list_voltage_measured,self.data_list_power,color='Red', linewidth=2, label='P-V curve')

            self.ax.legend(loc='upper right')

            self.canvas_chart.draw()

    # Function to setup chart
    def setup_chart(self) :

        self.fig_chart, self.ax = plt.subplots(figsize=(6.7, 4))
        self.fig_chart.patch.set_facecolor("white")
        self.fig_chart.patch.set_linewidth(2)

        self.configure_plot()

        self.canvas_chart = FigureCanvasTkAgg(self.fig_chart, master=self.dashboard_frame)
        self.canvas_chart.draw()
        self.canvas_chart.get_tk_widget().place(x=43, y=77)




    # Function to Start Test if RUN TEST button is pressed
    def run_test(self):
        self.reset_variables()
        self.running = True
        self.progress = 0
        self.progress_bar.set(self.progress)
        self.status_label.configure(text="Running...", text_color="orange")
        self.run_button.configure(state='disabled' ,image=self.test_running_img)


        if self.mode_var.get() == "CV":
            start_voltage = self.working_profile["start voltage"].get()
            stop_voltage = self.working_profile["stop voltage"].get()
            step_size_voltage = self.working_profile["step size voltage"].get()
            dwell_time_voltage = self.working_profile["dwell time voltage"].get()
            current_limit = self.working_profile["current limit"].get()
            self.plot_interval = dwell_time_voltage
            self.bk_device.set_CV(current_limit)

            self.data_list_voltage_measured.clear()
            self.data_list_current_measured.clear()
            self.data_list_power.clear()
            if step_size_voltage and stop_voltage and dwell_time_voltage:

                if start_voltage < stop_voltage :
                    voltages = np.arange(start_voltage, stop_voltage + step_size_voltage, step_size_voltage)
                    self.bk_device.set_voltage(voltages[0])
                    self.dashboard_frame.after(1000, self.process_next_voltage, voltages, 0)
                else :
                    self.running = False
                    self.progress_label.configure(text="100%")
                    self.status_label.configure(text="  Saved", text_color="#06F30B")
                    self.run_button.configure(state='normal',image=self.run_test_img)
                    messagebox.showinfo("Error", "Start voltage must be less than Stop voltage")
            else :
                self.running = False
                self.progress_label.configure(text="100%")
                self.status_label.configure(text="  Saved", text_color="#06F30B")
                self.run_button.configure(state='normal',image=self.run_test_img)
                messagebox.showinfo("Error", "Same Profile Data is missing")

        elif self.mode_var.get() == "CC":
            pass

    # Function to move to next voltage value (in order to accepte dwell time)
    def process_next_voltage(self,voltages,index):
        if not self.running  or index >= len(voltages) :
            self.running = False
            self.calculate_FF_Grade()
            self.calculate_recurrence()
            self.find_result()
            # self.Arduino.update_temperature()
            self.show_table()
            self.CollectData()
            self.progress_label.configure(text="100%")
            self.status_label.configure(text="  Saved", text_color="#06F30B")
            self.run_button.configure(state='normal',image=self.run_test_img)
            messagebox.showinfo("Information", "              Test Data Saved                ")

            return

        voltage = voltages[index]
        self.bk_device.set_voltage(voltage)

        # Schedule the next update
        self.dashboard_frame.after(int(self.plot_interval * 1000), lambda: self.update_data_and_plot(voltages,index))

    # Function to update date and change it in the chart
    def update_data_and_plot(self,voltages,index):
        if not self.running:
            return

        current_measured,voltage_measured = self.get_data()
        voltage = voltages[index]
        # current_measured = 10 * math.sin(2 * 0.1) + random.uniform(-1, 1)
        # voltage_measured = 20 * math.cos(2 * 0.1) + random.uniform(-1, 1)

        power = current_measured * voltage_measured
        

        self.Current_var.set(current_measured)
        self.Voltage_var.set(voltage_measured)
        self.Power_var.set(power)
        self.Current_formated.set(f"{self.Current_var.get():.2f}")
        self.Voltage_formated.set(f"{self.Voltage_var.get():.2f}")
        self.Power_formated.set(f"{self.Power_var.get():.2f}")
        
        self.data_list_voltage_measured.append(voltage_measured)
        self.data_list_current_measured.append(current_measured)
        self.data_list_power.append(power)

        self.update_data()

        self.progress += 1 / len(voltages)
        self.progress_bar.set(self.progress)
        self.progress_label.configure(text=f"{int(self.progress * 100)}%")

        self.animate_chart(current_measured, voltage_measured)

        self.process_next_voltage(voltages, index + 1)




    # Function to Get Data and send it to SaveData function That Add it to excel file
    def CollectData(self) :

        current_date = datetime.now()
        formatted_date = current_date.strftime("%m/%d/%Y")
        formatted_time = current_date.strftime("%H:%M:%S")

        serial_number = self.get_serialNum()
        Voc =self.Voc_var.get()
        Isc = self.Isc_var.get()
        max_power = self.max_power_var.get()
        Impp = self.Impp_var.get()
        Vmpp = self.Vmpp_var.get()
        FF = self.FF_var.get()
        grade = self.grade_var.get()
        result = self.result_var.get()
        recurrence = self.recurrence_var.get()

        self.SaveData(formatted_date,formatted_time,serial_number,max_power,Impp,Vmpp,Voc,Isc,FF,grade,result,recurrence)
    
    # Function to create excel file in not found
    def create_excel_file_if_not_exists(self):

        
        month_folder = self.get_month_folder()

        excel_file_path = os.path.join(month_folder, 'Data.xlsx')

        if not os.path.exists(excel_file_path):
            wb = Workbook()

            ws = wb.active
            ws.title = "Sheet1"

            headers = ["ID", "Module Number", "Test Result", "Pmpp", "Imp", "Vmp", "Uoc", "Isc",
                    "Fil factor", "Grade", "Recurrence", "Temperature Ambient", "Temperature Lamps",
                    "Pmpp Reference", "Pmpp Deviation", "Uoc Reference", "Uoc Deviation", "Isc Reference",
                    "Isc Deviation", "Temperature Ambient Reference", "Temperature Lamps Reference",
                    "Reference Number", "Date", "Time", "Serial Number"]
            ws.append(headers)
            wb.save(excel_file_path)

    # Function fo save data in the excel file
    def SaveData(self, date, time, serial_number, max_power=0, Impp=0, Vmpp=0, Voc=0, Isc=0, FF=0, Grade="A",result = "Passe",recurrence = 1):

        month_folder = self.get_month_folder()
        excel_file_path = os.path.join(month_folder, 'Data.xlsx')

        wb = load_workbook(excel_file_path)
        sheet = wb['Sheet1']

        last_id = sheet.cell(row=sheet.max_row, column=1).value
        last_module_number = sheet.cell(row=sheet.max_row, column=2).value 

        next_id = int(last_id) + 1 if last_id is not None and str(last_id).isdigit() else 1
        next_module_number = int(last_module_number) + 1 if last_module_number is not None and str(last_module_number).isdigit() else 312

        row = (next_id,                             # ID
            next_module_number,                     # Module Number
            result,                                 # Test Result
            max_power,                              # Pmpp
            Impp,                                   # Imp
            Vmpp,                                   # Vmp
            Voc,                                    # Uoc
            Isc,                                    # Isc
            FF,                                     # Fil factor
            Grade,                                  # Grade
            recurrence,                             # Recurrence
            0,                                      # Temperature Ambient
            0,                                      # Temperature Lamps
            31.0654,                                # Pmpp Reference
            ((max_power - 31.0654)/31.0654)*100 ,   # Pmpp Deviation
            9.9332,                                 # Uoc Reference
            ((Voc - 9.9332)/9.9332)*100,            # Uoc Deviation
            4.3456,                                 # Isc Reference
            ((Isc - 4.3456)/4.3456) * 100,          # Isc Deviation
            0,                                      # Temperature Ambient Reference
            2,                                      # Temperature Lamps Reference
            0,                                      # Reference Number
            date,                                   # Date
            time,                                   # Time
            serial_number                           # Serial Number
            )

        sheet.append(row)

        # Save the workbook
        wb.save(excel_file_path)

    # Function to Add data to table bellow the screen
    def Add_data_table(self) :

        current_date = datetime.now()
        formatted_date = current_date.strftime("%m/%d/%Y")
        formatted_time = current_date.strftime("%H:%M:%S")

        serial_number = self.get_serialNum()
        Voc =self.Voc_formated.get()
        Isc = self.Isc_formated.get()
        max_power = self.max_power_formated.get()
        Impp = self.Impp_formated.get()
        Vmpp = self.Vmpp_formated.get()
        FF = self.FF_formated.get()
        grade = self.grade_var.get()
        recurrence_count = self.recurrence_var.get()
        result = self.result_var.get()

        new_data = {
        "Serial Num": serial_number,
        "Results" : result,
        "Date": formatted_date,
        "Test Time": formatted_time,
        "Mpp": max_power,
        "Imp": Impp,
        "Vmp": Vmpp,
        "Isc": Isc,
        "Voc": Voc,
        "FF": FF,
        "Grade": grade,
        "Recurrence" : recurrence_count,
        "Profile" : self.working_profile["name"]
        }

        new_row = pd.DataFrame([new_data])
        self.TableData = pd.concat([self.TableData, new_row], ignore_index=True)

    #   Function to reset variabels after test finished
    def reset_variables(self) :
        self.Current_var.set(0.00)
        self.Voltage_var.set(0.00)
        self.Power_var.set(0.00)
        self.max_power_var.set(0.00)
        self.Vmpp_var.set(0.00)
        self.Impp_var.set(0.00)
        self.Isc_var.set(0.00)
        self.FF_var.set(0.00)
        self.temp_var.set(25)
        self.grade_var.set("?")
        self.recurrence_var.set(0)




    # Function to turn ON the light
    def ON_Lamps(self):
        self.Arduino.turn_light_on()
        self.ON_button.configure(image = self.lamps_on_img)
        self.OFF_button.configure(image = self.lamps_button_img)

    # Function to turn OFF the light
    def OFF_Lamps(self):
        self.Arduino.turn_light_off()
        self.ON_button.configure(image = self.lamps_button_img)
        self.OFF_button.configure(image = self.lamps_off_img)




    #Function to show table
    def show_table(self) :
        self.Add_data_table()

        self.tree.delete(*self.tree.get_children())

        for index, row in self.TableData.iterrows():
            self.tree.insert("", "end", values=(
                row["Serial Num"], row["Results"], row["Date"], row["Test Time"],
                row["Mpp"], row["Imp"], row["Vmp"], row["Isc"], row["Voc"],
                row["FF"], row["Grade"], row["Recurrence"],
                row["Profile"]
            ))
    
    # Function to show content of dashboard frame if DASHBOARD button is pressed
    def show_dashboard(self):
        self.dashboard_frame.pack(fill="both", expand=True)
        self.dashboard_button.configure(image = self.image_dashboard_ON_img, text_color = "#0000ff")
        self.bk_profiles_button.configure(image = self.image_bk_profiles_OFF_img, text_color = "#b7b7fc")
        self.bk_profiles_frame.pack_forget()

    # Function to show content of Bk Profiles frame if BK PROFILES button is pressed
    def show_bk_profiles(self):
        if self.prompt_for_password() :
            self.bk_profiles_frame.pack(fill="both", expand=True)
            self.dashboard_button.configure(image = self.image_dashboard_OFF_img, text_color = "#b7b7fc")
            self.bk_profiles_button.configure(image = self.image_bk_profiles_ON_img, text_color = "#0000ff")
            self.dashboard_frame.pack_forget()
        else :
            pass





    # Function to intialise new profile
    def initialize_new_profile(self):

        return {
            "name" : self.selected_option.get(),
            "start voltage": tk.DoubleVar(value=0),
            "stop voltage": tk.DoubleVar(value=0),
            "step size voltage": tk.DoubleVar(value=0),
            "dwell time voltage": tk.DoubleVar(value=0),
            "start current": tk.DoubleVar(value=0),
            "stop current": tk.DoubleVar(value=0),
            "step size current": tk.DoubleVar(value=0),
            "dwell time current": tk.DoubleVar(value=0),
            "current limit": tk.DoubleVar(value=0),
            "voltage limit": tk.DoubleVar(value=0),
            "power limit": tk.DoubleVar(value=0),
            "temperature limit": tk.DoubleVar(value=0),
            "current resolution": tk.DoubleVar(value=0),
            "voltage resolution": tk.DoubleVar(value=0),
            "active profile": 0,
        }

    # Function to change mode between CC,CV,CR and CP
    def mode_changed(self) :
        if self.mode_var.get() == "CV" :
            self.cv_radio.configure(fg_color = "#00ff00")
            self.bk_canvas.itemconfig(self.set_voltage_sweep, fill="#00A3FF")
            self.bk_canvas.itemconfig(self.set_current_sweep,fill = "black")

            self.bk_canvas.itemconfig(self.voltage_sweep_frame, image = self.voltage_sweep_frame_active_img)
            self.bk_canvas.itemconfig(self.current_sweep_frame, image = self.current_sweep_frame_img)

            self.bk_canvas.itemconfigure(self.entry_start_current_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_stop_current_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_step_size_current_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_dwell_time_current_unit, state="hidden")

            self.bk_canvas.itemconfigure(self.entry_start_voltage_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_stop_voltage_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_step_size_voltage_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_dwell_time_voltage_unit, state="normal")


            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_start_voltage, image = self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_stop_voltage, image = self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_step_size_voltage, image = self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_dwell_time_voltage, image = self.Bk_profilles_entry_data_active_img)

            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_start_current, image=self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_stop_current, image=self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_step_size_current, image=self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_dwell_time_current, image=self.Bk_profilles_entry_data_img)

            self.entry_start_voltage.place(x=723, y=466)
            self.entry_stop_voltage.place(x=723, y=506)
            self.entry_step_size_voltage.place(x=723, y=546)
            self.entry_dwell_time_voltage.place(x=723, y=586)


            self.entry_start_current.place_forget()
            self.entry_stop_current.place_forget()
            self.entry_step_size_current.place_forget()
            self.entry_dwell_time_current.place_forget()




        elif self.mode_var.get() == "CC" :
            self.cc_radio.configure(fg_color = "#00ff00")
            self.bk_canvas.itemconfig(self.set_voltage_sweep, fill="black")
            self.bk_canvas.itemconfig(self.set_current_sweep,fill = "#00A3FF")

            
            self.bk_canvas.itemconfig(self.voltage_sweep_frame, image = self.voltage_sweep_frame_img)
            self.bk_canvas.itemconfig(self.current_sweep_frame, image = self.current_sweep_frame_active_img)

            self.bk_canvas.itemconfigure(self.entry_start_current_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_stop_current_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_step_size_current_unit, state="normal")
            self.bk_canvas.itemconfigure(self.entry_dwell_time_current_unit, state="normal")

            self.bk_canvas.itemconfigure(self.entry_start_voltage_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_stop_voltage_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_step_size_voltage_unit, state="hidden")
            self.bk_canvas.itemconfigure(self.entry_dwell_time_voltage_unit, state="hidden")

            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_start_voltage, image = self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_stop_voltage, image = self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_step_size_voltage, image = self.Bk_profilles_entry_data_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_dwell_time_voltage, image = self.Bk_profilles_entry_data_img)

            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_start_current, image=self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_stop_current, image=self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_step_size_current, image=self.Bk_profilles_entry_data_active_img)
            self.bk_canvas.itemconfig(self.Bk_profiles_entry_data_dwell_time_current, image=self.Bk_profilles_entry_data_active_img)


            self.entry_start_voltage.place_forget()
            self.entry_stop_voltage.place_forget()
            self.entry_step_size_voltage.place_forget()
            self.entry_dwell_time_voltage.place_forget()

            self.entry_start_current.place(x=723, y=246)
            self.entry_stop_current.place(x=723, y=286)
            self.entry_step_size_current.place(x=723, y=326)
            self.entry_dwell_time_current.place(x=723, y=366)




        elif self.mode_var.get() == "CP" :
            self.cp_radio.configure(fg_color = "#00ff00") 


        elif self.mode_var.get() == "CR" :
            self.cr_radio.configure(fg_color = "#00ff00")
    
    # Function to state of entries and buttons depending on mode selected    
    def check_profile(self) :
        if self.selected_option.get() == "Profile 1" or self.selected_option.get() == "Profile 2" :
            self.entry_start_current.configure(state = "readonly")
            self.entry_stop_current.configure(state = "readonly")
            self.entry_step_size_current.configure(state = "readonly")
            self.entry_dwell_time_current.configure(state = "readonly")
            self.entry_start_voltage.configure(state = "readonly")
            self.entry_stop_voltage.configure(state = "readonly")
            self.entry_step_size_voltage.configure(state = "readonly")
            self.entry_dwell_time_voltage.configure(state = "readonly")
            self.entry_current_limit.configure(state = "readonly")
            self.entry_voltage_limit.configure(state = "readonly")
            self.entry_power_limit.configure(state = "readonly")
            self.entry_temperature_limit.configure(state = "readonly")
            self.entry_current_resolution.configure(state = "readonly")
            self.entry_voltage_resolution.configure(state = "readonly")

            self.save_profile_button.configure(state = "disabled",image = self.save_profile_disabled_img)
            self.delete_profile_button.configure(state = "disabled", image = self.delete_profile_disabled_img)
        else :
            self.entry_start_current.configure(state = "normal")
            self.entry_stop_current.configure(state = "normal")
            self.entry_step_size_current.configure(state = "normal")
            self.entry_dwell_time_current.configure(state = "normal")
            self.entry_start_voltage.configure(state = "normal")
            self.entry_stop_voltage.configure(state = "normal")
            self.entry_step_size_voltage.configure(state = "normal")
            self.entry_dwell_time_voltage.configure(state = "normal")
            self.entry_current_limit.configure(state = "normal")
            self.entry_voltage_limit.configure(state = "normal")
            self.entry_power_limit.configure(state = "normal")
            self.entry_temperature_limit.configure(state = "normal")
            self.entry_current_resolution.configure(state = "normal")
            self.entry_voltage_resolution.configure(state = "normal")   

            self.save_profile_button.configure(state = "normal",image = self.save_profile_img)
            self.delete_profile_button.configure(state = "normal", image = self.delete_profile_img)

    # Function to change profile
    def change_profile(self, selected_value):

        if self.selected_option.get().startswith("Profile") :
            self.selected_index = int(selected_value.split()[-1]) - 1
            self.selected_profile = self.profiles[self.selected_index]

            if self.selected_profile["active profile"] == 1 :

                self.activate_profile_button.configure(image = self.activate_profile_img,text = "Activated", text_color='#03FF0D')
                self.update_entries()
                self.working_profile = self.selected_profile
            else :
                self.selected_profile = self.profiles[self.selected_index]
                self.activate_profile_button.configure(image = self.disabled_profile_img, text = "Disabled", text_color='#FF0303')
                self.update_entries()
        else:
            self.activate_profile_button.configure(image = self.disabled_profile_img, text = "Disabled", text_color='#FF0303')
            self.new_profile = self.initialize_new_profile()
            self.selected_profile = self.new_profile.copy()
            self.update_entries()

                
        self.check_profile()

    # Function to save profile
    def save_profile(self) :

        last_profile_number = int(self.options_list[-2].split()[1])
        new_profile_name = "Profile " + str(last_profile_number+1)
        self.selected_profile["name"] = new_profile_name

        self.profiles.append(self.selected_profile)
        self.options_list.insert(-1, new_profile_name)
        self.option_menu.configure(values=self.options_list)

    # Function to delete profile Added
    def delete_profile(self) :
        if self.selected_option.get().startswith("Profile") :
            self.profiles.pop(self.selected_index)
            self.options_list.pop(self.options_list.index(self.selected_option.get()))

            self.selected_option.set(self.options_list[0])
            self.option_menu.configure(values = self.options_list)

            self.change_profile("Profile 1")

    # Function to make profile chosen as a working profile
    def activate_profile(self) :
        if self.selected_profile["active profile"] == 1:
            self.activate_profile_button.configure(image = self.disabled_profile_img, text = "Disabled", text_color='#FF0303')
            self.selected_profile["active profile"] = 0
            self.working_profile = self.profiles[0]
        else:
            self.activate_profile_button.configure(image = self.activate_profile_img,text = "Activated", text_color='#03FF0D')
            self.selected_profile["active profile"] = 1
            self.working_profile = self.selected_profile

    # Function to update entries if selected profile changed
    def update_entries(self):
        self.entry_start_current.configure(textvariable = self.selected_profile["start current"] )
        self.entry_stop_current.configure(textvariable = self.selected_profile["stop current"])
        self.entry_step_size_current.configure(textvariable = self.selected_profile["step size current"] )
        self.entry_dwell_time_current.configure(textvariable = self.selected_profile["dwell time current"] )
        self.entry_start_voltage.configure(textvariable = self.selected_profile["start voltage"] )
        self.entry_stop_voltage.configure(textvariable = self.selected_profile["stop voltage"] )
        self.entry_step_size_voltage.configure(textvariable = self.selected_profile["step size voltage"] )
        self.entry_dwell_time_voltage.configure(textvariable = self.selected_profile["dwell time voltage"] )
        self.entry_current_limit.configure(textvariable = self.selected_profile["current limit"] )
        self.entry_voltage_limit.configure(textvariable = self.selected_profile["voltage limit"] )
        self.entry_power_limit.configure(textvariable = self.selected_profile["power limit"] )
        self.entry_temperature_limit.configure(textvariable = self.selected_profile["temperature limit"] )
        self.entry_current_resolution.configure(textvariable = self.selected_profile["current resolution"] )
        self.entry_voltage_resolution.configure(textvariable = self.selected_profile["voltage resolution"] )




    # Function to validate serial number entred (6 numbers & containe only numbers)
    def validate_serial(self, new_value):
        if new_value.isdigit():
            if len(new_value) < 6:
                self.run_button.configure(state="disabled", image=self.run_test_disabled_img)

            elif len(new_value) == 6:
                self.run_button.configure(state="normal", image=self.run_test_img)

            return True

        elif new_value == "":
            return True

        else :
            self.run_button.configure(state="disabled", image=self.run_test_disabled_img)
            return False

    # Function to takes user password
    def prompt_for_password(self):
        password = simpledialog.askstring("Password", "\n\n\t\t\tEnter password:\t\t\t\t\n")
        if password == "agamine12":
            return True
        else:
            return False

    # Function to display infos about the application
    def About(self) :
        messagebox.showinfo("About", "PV Module Testing Application \n\nDesigned by Mohamed EL Hamdani\n     Email : service4@agamine.com \n\nMade by Abdelmajid Sabiri\n      Email : abdelmajidsabiri77@gmail.com\n      University : EST Essaouira\n\n Date of creation : 08/2024")


    # Funtion to Setup content of Dashboard Frame
    def setup_dashboard_content(self):
        
        # initialis dashborad canvas
        canvas = Canvas(
            self.dashboard_frame,
            bg = "#D7E1E7",
            height = 990,
            width = 1620,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )

        # Add images and Text
        canvas.place(x = 0, y = 0)
        self.run_test_img = PhotoImage(file = self.get_image_path("run_test.png"))
        self.run_test_disabled_img = PhotoImage(file=self.get_image_path("run_test_disabled.png"))
        self.test_running_img = PhotoImage(file = self.get_image_path("test_running.png"))
        self.insert_SN_img = PhotoImage(file=self.get_image_path("insert_SN.png"))
        self.plot_img = PhotoImage(file = self.get_image_path("plot_background.png"))
        self.data_img = PhotoImage(file = self.get_image_path("data_background.png"))
        self.vol_curr_pow_img = PhotoImage(file = self.get_image_path("vol_curr_pow_background.png"))
        self.lamps_img = PhotoImage(file = self.get_image_path("lamps_background.png"))
        self.table_img = PhotoImage(file = self.get_image_path("table_background.png"))
        self.progress_bar_img = PhotoImage(file = self.get_image_path("progress_bar_background.png"))
        self.status_img = PhotoImage(file = self.get_image_path("status.png"))
        self.vol_curr_pow_values_img = PhotoImage(file = self.get_image_path("vol_curr_pow_values_background.png"))
        self.data_values_img = PhotoImage(file = self.get_image_path("data_values_background.png"))
        self.line_data_img = PhotoImage(file = self.get_image_path("line.png"))
        self.lamps_on_img = PhotoImage(file = self.get_image_path("lamps_on.png"))
        self.lamps_off_img = PhotoImage(file = self.get_image_path("lamps_off.png"))
        self.lamps_button_img = PhotoImage(file = self.get_image_path("lamps_button.png"))



        insert_SN_background = canvas.create_image(           250,50,image=self.insert_SN_img)
        plot_background = canvas.create_image(     380,280,image = self.plot_img)
        data_background = canvas.create_image(     1020,280,image = self.data_img)
        vol_curr_pow_background = canvas.create_image(     380,535,image = self.vol_curr_pow_img)
        lamp_ON_background = canvas.create_image(  900,535,image = self.lamps_img)
        lamp_OFF_background = canvas.create_image( 1140,535,image = self.lamps_img)
        table_background = canvas.create_image(    642,770,image = self.table_img)

        fram_indicator = canvas.create_image(      4,45.0,image=self.fram_indicator_img)
        vertical_line = canvas.create_image(       0,400,image=self.vertical_line_img)
        progress_bar_background = canvas.create_image(        760, 48, image=self.progress_bar_img )
        test_status_background = canvas.create_image(1100,30, image = self.status_img)
        line_data = canvas.create_image(970, 102, image = self.line_data_img)
        line_data = canvas.create_image(1090, 102, image = self.line_data_img)

        isc_value_background = canvas.create_image(1030, 140, image = self.data_values_img)
        Voc_value_background = canvas.create_image(1030, 185, image = self.data_values_img)
        Mpp_value_background = canvas.create_image(1030, 230, image = self.data_values_img)
        Ipm_value_background = canvas.create_image(1030, 275, image = self.data_values_img)
        Vpm_value_background = canvas.create_image(1030, 320, image = self.data_values_img)
        FF_value_background = canvas.create_image(1030, 365, image = self.data_values_img)
        Grade_value_background = canvas.create_image(1030, 410, image = self.data_values_img)
        test_recurrence_background = canvas.create_image(1030, 455, image = self.data_values_img)

        Vol_background = canvas.create_image(170,535, image = self.vol_curr_pow_values_img)
        Curr_background = canvas.create_image(410,535, image = self.vol_curr_pow_values_img)
        Pow_background = canvas.create_image(640,535, image = self.vol_curr_pow_values_img)


        


        canvas.create_text(25,20, anchor="nw", text="Insert SN", fill="black", font=("",13))
        canvas.create_text(950,20, anchor="nw", text="Test status", fill="black", font=("",13))
        canvas.create_text(850,92, anchor="nw", text="Description", fill="#9ea5d2", font=("Helvetica",10, "bold"))
        canvas.create_text(1010,92, anchor="nw", text="Value", fill="#9ea5d2", font=("Helvetica",10, "bold"))
        canvas.create_text(1130,92, anchor="nw", text="Unit", fill="#9ea5d2", font=("Helvetica",10, "bold"))


        canvas.create_text(945,130, anchor="nw", text="Isc", fill="#0000FF", font=("",9))
        canvas.create_text(945,175, anchor="nw", text="Voc", fill="#0000FF", font=("",9))
        canvas.create_text(945,220, anchor="nw", text="Mpp", fill="#0000FF", font=("",9))
        canvas.create_text(945,265, anchor="nw", text="Ipm", fill="#0000FF", font=("",9))
        canvas.create_text(945,310, anchor="nw", text="Vpm", fill="#0000FF", font=("",9))
        canvas.create_text(945,355, anchor="nw", text="FF", fill="#0000FF", font=("",9))
        canvas.create_text(945,400, anchor="nw", text="G", fill="#0000FF", font=("",9))

        canvas.create_text(830,130, anchor="nw", text="Short Circuit Current", fill="Black", font=("",9))
        canvas.create_text(830,175, anchor="nw", text="Open Circuit Voltage", fill="Black", font=("",9))
        canvas.create_text(820,220, anchor="nw", text="Maximum Power Point", fill="Black", font=("",9))
        canvas.create_text(860,265, anchor="nw", text="Current at MPP", fill="Black", font=("",9))
        canvas.create_text(860,310, anchor="nw", text="Voltage at MPP", fill="Black", font=("",9))
        canvas.create_text(890,355, anchor="nw", text="Fil Factor", fill="Black", font=("",9))
        canvas.create_text(910,400, anchor="nw", text="Grade", fill="Black", font=("",9))
        canvas.create_text(880,445, anchor="nw", text="Test Recurrence", fill="Black", font=("",9))

        canvas.create_text(1140,130, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,175, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,220, anchor="nw", text="W", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,265, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,310, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,355, anchor="nw", text="%", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,400, anchor="nw", text="⭐", fill="#0000FF", font=("Helvetica",10, "bold"))
        canvas.create_text(1140,445, anchor="nw", text="#", fill="#0000FF", font=("Helvetica",10, "bold"))


        canvas.create_text(60,527, anchor="nw", text="Voltage", fill="Black", font=("",11))
        canvas.create_text(300,527, anchor="nw", text="Current", fill="Black", font=("",11))
        canvas.create_text(538,527, anchor="nw", text="Power", fill="Black", font=("",11))

        canvas.create_text(200,527, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11, "bold"))
        canvas.create_text(440,527, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11, "bold"))
        canvas.create_text(665,527, anchor="nw", text="W", fill="#0000FF", font=("Helvetica",11, "bold"))

        canvas.create_text(962,527, anchor="nw", text="On", fill="#0000FF", font=("Helvetica",11, "bold"))
        canvas.create_text(1197,527, anchor="nw", text="Off", fill="#0000FF", font=("Helvetica",11, "bold"))


        self.setup_chart()

        # Entry Text for serial number of solar module
        self.entry_serialNum = ctk.CTkEntry(master=self.dashboard_frame, placeholder_text="Serial Number",textvariable=self.serial_num_var ,border_width = 0, fg_color = "#f6f6f6", bg_color="#f6f6f6", width=230,validate="key",validatecommand=(self.dashboard_frame.register(self.validate_serial), "%P"),font=("Helvetica",14, "bold"))
        self.entry_serialNum.place(x=120, y=17)

        # RUN TEST button to start test
        self.run_button = ctk.CTkButton(master=self.dashboard_frame,text = "RUN TEST",image=self.run_test_disabled_img, command=self.run_test, fg_color='#d2dce2', text_color='#0000FF', font=("Arial Rounded MT Bold",13),hover="transparent",state='disabled')
        self.run_button.place(x=480, y=6)


        # Progress bar to show the progress of the test
        self.progress_bar = ctk.CTkProgressBar(master=self.dashboard_frame, width = 150,height=8.5, bg_color= "white",progress_color = ("#00d9ff", "black"))
        self.progress_bar.set(self.progress)
        self.progress_bar.place(x=660, y=26)

        # Progress label to display the percentage
        self.progress_label = ctk.CTkLabel(master=self.dashboard_frame, text="0%", font=("Arial", 13, "bold"), text_color='#2021fd', bg_color="white")
        self.progress_label.place(x=820,y=14)

        # Status Label to display the state of the test
        self.status_label = ctk.CTkLabel(master=self.dashboard_frame, text="Waiting", text_color="orange", font=("Arial", 12, "bold"), bg_color="white")
        self.status_label.place(x=1077, y=17)

 

        self.Isc_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Isc_formated, height=5 , width=20 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.Isc_label.place(x=1010, y=132)

        self.Voc_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Voc_formated, height=5 , width=20 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.Voc_label.place(x=1010, y=177)

        self.Mpp_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.max_power_formated, height=5 , width=20 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.Mpp_label.place(x=1010, y=222)

        self.Impp_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Impp_formated, height=5 , width=20 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.Impp_label.place(x=1010, y=267)

        self.Vmpp_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Vmpp_formated, height=5 , width=20 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.Vmpp_label.place(x=1010, y=312)
        
        self.FF_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.FF_formated, height=5 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.FF_label.place(x=1010, y=357)

        self.grade_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.grade_var, height=5 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.grade_label.place(x=1010, y=402)

        # self.temp_label = ctk.CTkLabel(master = self.dashboard_frame, text = f"{round(self.temp_var.get(),2)}", height=5 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        # self.temp_label.place(x=1003, y=400)

        self.recurrence_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.recurrence_var, height=5 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 11.5,"bold"))
        self.recurrence_label.place(x=1010, y=447)

        self.Voltage_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Voltage_formated, height=5, width=60 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 13,"bold"))
        self.Voltage_label.place(x=133, y=527)

        self.Current_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Current_formated, height=5 ,width=60 ,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 13,"bold"))
        self.Current_label.place(x=373, y=527)

        self.Power_label = ctk.CTkLabel(master = self.dashboard_frame, textvariable = self.Power_formated, height=5, width=60,text_color="Black",bg_color = "#EBECF0" ,font=("Arial", 13,"bold"))
        self.Power_label.place(x=605, y=527)


        style = ttk.Style()
        style.configure("Custom.Treeview",
                        rowheight=50,
                        borderwidth=0, 
                        background="#D7E1E7",
                        foreground="#000000",
                        font=("Helvetica",9,"normal"))

        style.configure("Custom.Treeview.Heading",
                        borderwidth=0,  
                        relief="flat",
                        background="#C9D6DC",  
                        foreground="#000000", 
                        font=("Helvetica",10,"bold")) 

        # Configure selected item style
        style.map("Custom.Treeview",
                background=[('selected', '#D0EBF7')],
                foreground=[('selected', '#0000FF')])

        self.tree = ttk.Treeview(self.dashboard_frame, columns=self.columns, show="headings", height=2, style="Custom.Treeview")
        self.tree.place(x=40, y=600, width=1200, height=280)
          # Adjust width as needed
        for col in self.columns:
            self.tree.column(col, anchor="center",width = 50)
            self.tree.heading(col, text=col)


        self.vsb = ctk.CTkScrollbar(self.tree, orientation="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")


    


        # Lamps Button
        self.ON_button = ctk.CTkButton(master = self.dashboard_frame, 
                                        text="Turn Lamps", 
                                        image = self.lamps_button_img, 
                                        text_color="Black", 
                                        fg_color='white', 
                                        command= self.ON_Lamps, 
                                        font=("Arial",14), 
                                        compound = "right",
                                        hover_color="white",
                                        bg_color = "white")
        self.ON_button.place(x=814,y=510)

        self.OFF_button = ctk.CTkButton(master = self.dashboard_frame, 
                                        text="Turn Lamps", 
                                        image = self.lamps_off_img, 
                                        text_color="Black", 
                                        fg_color='white', 
                                        compound = "right", 
                                        command= self.OFF_Lamps,  
                                        font=("Arial",14),
                                        hover_color="white",
                                        bg_color="white")
        self.OFF_button.place(x=1050,y=510)

    # Function to Setup content of bk_profiles Frame
    def setup_bk_profiles_content(self):

        self.bk_canvas = Canvas(
            self.bk_profiles_frame,
            bg = "#D7E1E7",
            height = 930,
            width = 1721,
            bd = 0,
            highlightthickness = 0,
            relief = "ridge"
        )
        self.bk_canvas.place(x = 0, y = 0)


        self.Bk_profilles_frame_img = PhotoImage(file = self.get_image_path("bk_profiles_frame.png"))
        self.Bk_profilles_frame_top_img = PhotoImage(file = self.get_image_path("bk_profiles_frame_top.png"))
        self.Bk_profilles_entry_data_img = PhotoImage(file = self.get_image_path("bk_profiles_entry_data.png"))
        self.Bk_profilles_entry_data_active_img = PhotoImage(file = self.get_image_path("bk_profiles_entry_data_active.png"))
        self.safety_data_frame_img = PhotoImage(file = self.get_image_path("safety_data_frame.png"))
        self.set_measurement_frame_img = PhotoImage(file = self.get_image_path("set_measurement_precision_frame.png"))
        self.voltage_sweep_frame_img = PhotoImage(file =self.get_image_path("voltage_sweep_frame.png"))
        self.voltage_sweep_frame_active_img = PhotoImage(file = self.get_image_path("voltage_sweep_frame_active.png"))
        self.current_sweep_frame_img = PhotoImage(file = self.get_image_path("current_sweep_frame.png"))
        self.current_sweep_frame_active_img = PhotoImage(file = self.get_image_path("current_sweep_frame_active.png"))       
        self.mode_selection_frame_img = PhotoImage(file = self.get_image_path("mode_selection_frame.png"))
        self.save_profile_img = PhotoImage(file = self.get_image_path("save_profile.png"))
        self.save_profile_disabled_img = PhotoImage(file = self.get_image_path("save_profile_disabled.png"))
        self.delete_profile_img = PhotoImage(file = self.get_image_path("delete_profile.png"))
        self.delete_profile_disabled_img = PhotoImage(file = self.get_image_path("delete_profile_disabled.png"))
        self.activate_profile_img = PhotoImage(file=self.get_image_path("activate_profile.png"))
        self.disabled_profile_img = PhotoImage(file = self.get_image_path("disabled_profile.png"))

        





        fram_indicator = self.bk_canvas.create_image(4,148,image=self.fram_indicator_img)
        vertical_line = self.bk_canvas.create_image(0,400,image=self.vertical_line_img)
        Bk_profiles_frame = self.bk_canvas.create_image(630,400,image=self.Bk_profilles_frame_img)
        Bk_profiles_frame_top = self.bk_canvas.create_image(630,150,image=self.Bk_profilles_frame_top_img)
        mode_selection_frame = self.bk_canvas.create_image(255,320,image=self.mode_selection_frame_img)
        set_measurement_frame = self.bk_canvas.create_image(255,540,image=self.set_measurement_frame_img)        
        self.current_sweep_frame = self.bk_canvas.create_image(640,320,image=self.current_sweep_frame_img)
        self.voltage_sweep_frame = self.bk_canvas.create_image(640,540,image=self.voltage_sweep_frame_active_img)
        safety_data_frame = self.bk_canvas.create_image(1030,360,image=self.safety_data_frame_img)




        self.Bk_profiles_entry_data_start_current = self.bk_canvas.create_image(750,260,image = self.Bk_profilles_entry_data_img)
        self.Bk_profiles_entry_data_stop_current = self.bk_canvas.create_image(750,300,image = self.Bk_profilles_entry_data_img)
        self.Bk_profiles_entry_data_step_size_current = self.bk_canvas.create_image(750,340,image = self.Bk_profilles_entry_data_img)
        self.Bk_profiles_entry_data_dwell_time_current = self.bk_canvas.create_image(750,380,image = self.Bk_profilles_entry_data_img)

        self.Bk_profiles_entry_data_start_voltage = self.bk_canvas.create_image(750,480,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_stop_voltage = self.bk_canvas.create_image(750,520,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_step_size_voltage = self.bk_canvas.create_image(750,560,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_dwell_time_voltage = self.bk_canvas.create_image(750,600,image = self.Bk_profilles_entry_data_active_img)


        self.Bk_profiles_entry_data_current_resolution = self.bk_canvas.create_image(350,515,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_voltage_resolution = self.bk_canvas.create_image(350,555,image = self.Bk_profilles_entry_data_active_img)

        
        self.Bk_profiles_entry_data_current_limit = self.bk_canvas.create_image(1130,315,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_voltage_limit = self.bk_canvas.create_image(1130,355,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_power_limit = self.bk_canvas.create_image(1130,395,image = self.Bk_profilles_entry_data_active_img)
        self.Bk_profiles_entry_data_temperature_limit = self.bk_canvas.create_image(1130,435,image = self.Bk_profilles_entry_data_active_img)



        self.bk_canvas.create_text(60,55, anchor="nw", text="BK Profile Settings", fill="black", font=("Helvetica",16,"bold"))
        self.bk_canvas.create_text(675,140, anchor="nw", text="Profile", fill="black", font=("Helvetica",11,"bold"))
        self.bk_canvas.create_text(100,140, anchor="nw", text="Set Profile", fill="black", font=("Helvetica",11,"bold"))




        self.set_current_sweep = self.bk_canvas.create_text(500,230, anchor="nw", text="Set\nCurrent Sweep", fill="black", font=("Helvetica",10,"bold"))
        self.set_voltage_sweep = self.bk_canvas.create_text(500,450, anchor="nw", text="Set\nVoltage Sweep", fill="#00A3FF", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(112,450, anchor="nw", text="Set\nMeasurement Precision", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(885,230, anchor="nw", text="Set\nSafety and Protection", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(105,260, anchor="nw", text="Mode Selection", fill="black", font=("Helvetica",10,"bold"))

        self.bk_canvas.create_text(615,250, anchor="nw", text="Start Current", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,290, anchor="nw", text="Stop Current", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,330, anchor="nw", text="Step Size", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,370, anchor="nw", text="Dwell Time", fill="black", font=("Helvetica",10,"bold"))

        self.bk_canvas.create_text(615,470, anchor="nw", text="Start Voltage", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,510, anchor="nw", text="Stop Voltage", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,550, anchor="nw", text="Step Size", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(615,590, anchor="nw", text="Dwell Time", fill="black", font=("Helvetica",10,"bold"))

        self.bk_canvas.create_text(960,305, anchor="nw", text="Current Limit", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(960,345, anchor="nw", text="Voltage Limit", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(960,385, anchor="nw", text="Power Limit", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(960,425, anchor="nw", text="Temperature Limit", fill="black", font=("Helvetica",10,"bold"))

        self.bk_canvas.create_text(180,504, anchor="nw", text="Current Resolution", fill="black", font=("Helvetica",10,"bold"))
        self.bk_canvas.create_text(180,544, anchor="nw", text="Voltage Resolution", fill="black", font=("Helvetica",10,"bold"))

    
        self.option_menu = ctk.CTkOptionMenu(self.bk_profiles_frame, 
                                            values=self.options_list, 
                                            variable=self.selected_option,
                                            command = self.change_profile,
                                            corner_radius=20,
                                            fg_color = "#f6f7f9",
                                            bg_color= "#BBBBBB",
                                            text_color = "black",
                                            font = ("Helvetica",11,"bold"),
                                            button_color = "#f6f7f9",
                                            button_hover_color = "#E9E9E9")
        self.option_menu.place(x=185,y=136)

        self.cv_radio = ctk.CTkRadioButton(
            self.bk_profiles_frame,
            text="Constant Voltage CV",
            font = ("Helvetica",11,"bold"),
            fg_color = "#00ff00",
            bg_color="#F8F9FA",
            variable=self.mode_var,
            value="CV",
            command=self.mode_changed,
            radiobutton_width=20,
            radiobutton_height=20,
            corner_radius=5, 
            border_color = "Black",
            border_width_unchecked=1.5,
            )
        self.cv_radio.place(x=230,y=260)
        
        self.cc_radio = ctk.CTkRadioButton(
            self.bk_profiles_frame,
            text="Constant Current CC",
            font = ("Helvetica",11,"bold"),
            bg_color="#F8F9FA",
            variable=self.mode_var,
            value="CC",
            command=self.mode_changed,
            radiobutton_width=20,
            radiobutton_height=20,
            corner_radius=5, 
            border_color="Black",
            border_width_unchecked=1.5,
        )
        self.cc_radio.place(x=230,y=290)
        
        self.cp_radio = ctk.CTkRadioButton(
            self.bk_profiles_frame,
            text="Constant Power CP",
            font = ("Helvetica",11,"bold"),
            bg_color="#F8F9FA",
            variable=self.mode_var,
            value="CP",
            command=self.mode_changed,
            radiobutton_width=20, 
            radiobutton_height=20,
            corner_radius=5,
            border_color = "Black",
            border_width_unchecked=1.5,
        )
        self.cp_radio.place(x=230,y=320)
        
        self.cr_radio = ctk.CTkRadioButton(
            self.bk_profiles_frame,
            text="Constant Resistance CR",
            font = ("Helvetica",11,"bold"),
            bg_color="#F8F9FA",
            variable=self.mode_var,
            value="CR",
            command=self.mode_changed,
            radiobutton_width=20,
            radiobutton_height=20,
            corner_radius=5,
            border_color = "Black",
            border_width_unchecked=1.5,
        )
        self.cr_radio.place(x=230,y=350)

        self.save_profile_button = ctk.CTkButton(master=self.bk_profiles_frame, text = "Save Profile" ,image=self.save_profile_img, command=self.save_profile, fg_color='#BBBBBB',bg_color="#BBBBBB", text_color='black', font=("Helvetica",14, "bold"),hover="transparent",compound="right",text_color_disabled = "#999999")
        self.delete_profile_button = ctk.CTkButton(master=self.bk_profiles_frame, text = "Delete Profile" ,image=self.delete_profile_img, command=self.delete_profile, fg_color='#BBBBBB',bg_color="#BBBBBB", text_color='black', font=("Helvetica",14, "bold"),hover="transparent",compound="right",text_color_disabled = "#999999")
        self.activate_profile_button = ctk.CTkButton(master=self.bk_profiles_frame, text = "Activated" ,image=self.activate_profile_img, command=self.activate_profile, fg_color='#BBBBBB',bg_color="#BBBBBB", text_color='#03FF0D', font=("Helvetica",15, "bold"),hover="transparent",compound="right")
        self.save_profile_button.place(x=395, y=120)
        self.delete_profile_button.place(x=1000, y=124)
        self.activate_profile_button.place(x=723, y=129)

        

        self.entry_start_current = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0 ,textvariable= self.selected_profile["start current"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"))
        self.entry_stop_current = ctk.CTkEntry(master=self.bk_profiles_frame,border_width =0, textvariable= self.selected_profile["stop current"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"))
        self.entry_step_size_current = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["step size current"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"))
        self.entry_dwell_time_current = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["dwell time current"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"))

        self.entry_start_current_unit = self.bk_canvas.create_text(773,250, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11, "bold"),state="hidden")
        self.entry_stop_current_unit = self.bk_canvas.create_text(773,290, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11,"bold"),state="hidden")
        self.entry_step_size_current_unit = self.bk_canvas.create_text(773, 330, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11,"bold"),state="hidden")
        self.entry_dwell_time_current_unit = self.bk_canvas.create_text(773, 370, anchor="nw", text="S", fill="#0000FF", font=("Helvetica",11,"bold"),state="hidden")


        self.entry_start_voltage = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["start voltage"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_stop_voltage = ctk.CTkEntry(master=self.bk_profiles_frame,border_width =0,textvariable= self.selected_profile["stop voltage"], fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_step_size_voltage = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0,textvariable= self.selected_profile["step size voltage"], fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_dwell_time_voltage = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["dwell time voltage"],fg_color = "#D1FCFF", bg_color="#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_start_voltage.place(x=723, y=466)
        self.entry_stop_voltage.place(x=723, y=506)
        self.entry_step_size_voltage.place(x=723, y=546)
        self.entry_dwell_time_voltage.place(x=723, y=586)

        self.entry_start_voltage_unit = self.bk_canvas.create_text(773,472, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11, "bold"))
        self.entry_stop_voltage_unit = self.bk_canvas.create_text(773,512, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11,"bold"))
        self.entry_step_size_voltage_unit = self.bk_canvas.create_text(773, 552, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11,"bold"))
        self.entry_dwell_time_voltage_unit = self.bk_canvas.create_text(773, 592, anchor="nw", text="S", fill="#0000FF", font=("Helvetica",11,"bold"))


        self.entry_current_limit = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["current limit"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_voltage_limit = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["voltage limit"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_power_limit = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["power limit"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_temperature_limit = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["temperature limit"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_current_limit.place(x=1103, y=301)
        self.entry_voltage_limit.place(x=1103, y=341)
        self.entry_power_limit.place(x=1103, y=381)
        self.entry_temperature_limit.place(x=1103, y=421)

        self.entry_current_limit_unit = self.bk_canvas.create_text(1153,305, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11, "bold"))
        self.entry_voltage_limit_unit = self.bk_canvas.create_text(1153,345, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11,"bold"))
        self.entry_power_limit_unit = self.bk_canvas.create_text(1152, 385, anchor="nw", text="W", fill="#0000FF", font=("Helvetica",11,"bold"))
        self.entry_temperature_limit_unit = self.bk_canvas.create_text(1151.5, 425, anchor="nw", text="°C", fill="#0000FF", font=("Helvetica",11,"bold"))



        self.entry_current_resolution = ctk.CTkEntry(master=self.bk_profiles_frame,border_width =0, textvariable= self.selected_profile["current resolution"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_voltage_resolution = ctk.CTkEntry(master=self.bk_profiles_frame,border_width = 0, textvariable= self.selected_profile["voltage resolution"],fg_color = "#D1FCFF", bg_color = "#D1FCFF", width=50, text_color = "#0000FF",font = ("Helvetica",13,"bold"),)
        self.entry_current_resolution.place(x=323, y=501)
        self.entry_voltage_resolution.place(x=323, y=541)

        self.entry_current_resolution_unit = self.bk_canvas.create_text(374, 505, anchor="nw", text="A", fill="#0000FF", font=("Helvetica",11,"bold"))
        self.entry_voltage_resolution_unit = self.bk_canvas.create_text(374, 545, anchor="nw", text="V", fill="#0000FF", font=("Helvetica",11,"bold"))


        self.check_profile()

    # Funtion to close the application correctly
    def on_closing(self):
        self.bk_device.reset_to_manual()
        self.Arduino.close()
        self.window.destroy()
        self.window.quit()



# Instantiate the GUI class to start the application
if __name__ == "__main__":
    GUI()